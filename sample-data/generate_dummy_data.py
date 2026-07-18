"""Generate a realistic dummy research dataset that matches the real export
schema. Reuses app.scoring so the derived sheets are identical to production.

Participants play games across multiple days/sessions; some complete every
skill, some drop out, some never touch certain tasks — to show the missingness
and shape you'll actually analyse.

ALL VALUES ARE SYNTHETIC / FAKE. This is a format preview only.

Run (from the repo root, with the server venv + openpyxl):
    server/.venv/bin/python -m pip install openpyxl
    PYTHONPATH=server server/.venv/bin/python sample-data/generate_dummy_data.py

Outputs the workbook + one CSV per sheet next to this script.
"""
import csv
import os
import random
import uuid
from datetime import date, datetime, timedelta
from types import SimpleNamespace

from app import scoring
from app.scoring import age_band, age_years, iq_band

random.seed(20260718)

OUT = os.path.dirname(os.path.abspath(__file__))
os.makedirs(OUT, exist_ok=True)

# --- participants -------------------------------------------------------------
# (code, gender, dob, autism_level, iq, profile)
PARTICIPANTS = [
    ("P-001", "M", date(2016, 3, 12), "Level 1", 96, "completer"),
    ("P-002", "F", date(2013, 8, 2), "Level 1", 104, "completer"),
    ("P-003", "M", date(2017, 1, 25), "Level 2", 78, "partial"),
    ("P-004", "M", date(2011, 11, 9), "Level 2", 86, "completer"),
    ("P-005", "F", date(2018, 6, 30), "Level 3", 61, "dropout"),
    ("P-006", "M", date(2015, 4, 18), "Level 2", 89, "completer"),
    ("P-007", "F", date(2012, 9, 5), "Level 1", 109, "completer"),
    ("P-008", "M", date(2016, 12, 1), "Level 3", 70, "partial"),
]

# Which games each profile touches (skill coverage varies on purpose).
EMOTION = ["emotionrecognition", "identifyemotions", "emotionrecognition360"]
TURN = ["blocks", "rollback", "football360"]
NORMS = ["rightway", "rulefixer"]
JA = ["museum", "discovery", "museum360"]
VR_GAMES = {"emotionrecognition360", "identifyemotions360", "playroom360",
            "football360", "museum360", "park360"}  # xrPresenting=1 + head telemetry

PROFILE_GAMES = {
    "completer": EMOTION + TURN + NORMS + JA,          # all four skills
    "partial": EMOTION + TURN + JA,                    # skips social norms
    "dropout": ["emotionrecognition", "museum"],       # a couple, then stops
}
PROFILE_SESSIONS = {"completer": (3, 5), "partial": (2, 3), "dropout": (1, 2)}

STUDY_START = datetime(2026, 6, 1, 15, 30)
LEVELS = ["easy", "medium", "hard"]
QUIZ_CHANCE = {"easy": 0.5, "medium": 1 / 3, "hard": 1 / 3}
CLIPS_CHANCE = {"easy": 0.5, "medium": 1 / 3, "hard": 0.25}
CUES = ["verbal", "gesture", "orient"]
MUSEUM_CUES = [("pulse", "gesture", 3), ("hover", "gesture", 4), ("distal", "gaze", 6)]
EMOTIONS = ["happy", "sad", "angry", "surprised", "scared", "disgust"]
CONSTRUCTS = {
    "rightway": ["greetings", "sharing", "turns", "space", "politeness"],
    "rulefixer": ["helping", "comforting", "inclusion", "politeness", "fairness"],
    "rightway360": ["greetings", "sharing", "turns", "space", "politeness"],
}


def clamp(x, lo=0.02, hi=0.99):
    return max(lo, min(hi, x))


def head_block(target_bearing):
    """A plausible VR head-scan summary for one trial."""
    travel = round(abs(target_bearing) + random.uniform(10, 60), 1)
    return {
        "headSamples": random.randint(20, 70),
        "headStartYawDeg": round(random.uniform(-15, 15), 1),
        "headEndYawDeg": round(target_bearing + random.uniform(-6, 6), 1),
        "headYawTravelDeg": travel,
        "headYawRangeDeg": round(abs(target_bearing) + random.uniform(5, 25), 1),
        "headReversals": random.randint(0, 4),
        "headMaxPitchDeg": round(random.uniform(2, 18), 1),
        "headMinPitchDeg": round(random.uniform(-18, -2), 1),
        "headToTargetMs": random.randint(350, 2200),
    }


class Ev(SimpleNamespace):
    pass


def make_events(game, sess_id, student_id, user_id, t0, acc, lat):
    """Yield realistic events for one game session. `acc` = first-attempt success
    prob for this session; `lat` = median latency ms. Returns (events, final_score)."""
    evs = []
    step = 0
    vr = game in VR_GAMES

    def add(etype, payload, ev_score=None, dt_ms=None):
        nonlocal step
        base = {"xrPresenting": vr}
        base.update(payload)
        evs.append(Ev(
            id=uuid.uuid4(), student_id=student_id, user_id=user_id, session_id=sess_id,
            game_key=game, event_type=etype, step_index=step, score=ev_score,
            payload=base, client_timestamp=None,
            created_at=t0 + timedelta(milliseconds=(dt_ms if dt_ms is not None else step * 4200)),
        ))
        step += 1

    score = 0
    n = random.randint(5, 9)

    if game in ("emotionrecognition", "emotionrecognition360"):
        for k in range(n):
            level = random.choice(LEVELS)
            correct = random.random() < acc
            score += 1 if correct else 0
            bearing = random.choice([-70, -40, 40, 70]) if vr else 0
            p = {"correct": correct, "answer": random.choice(EMOTIONS),
                 "picked": random.choice(EMOTIONS), "kind": "whoFeels",
                 "latencyMs": max(500, int(random.gauss(lat, 350)))}
            if vr:
                p.update({"chance": round(QUIZ_CHANCE[level], 4), "hinted": random.random() < 0.25,
                          "targetBearingDeg": bearing, "boardCount": 2 if level == "easy" else 3,
                          "latencyFromPromptEndMs": max(250, int(random.gauss(lat - 700, 300)))})
                p.update(head_block(bearing))
            else:
                p["level"] = level
            add("answer", p, ev_score=score)

    elif game in ("identifyemotions", "identifyemotions360"):
        for k in range(n):
            level = random.choice(LEVELS)
            first_ok = random.random() < acc
            key = "difficulty" if vr else "level"
            common = {"clip": f"clip{random.randint(1, 12)}", key: level,
                      "answer": random.choice(EMOTIONS), "freezeKind": "peak"}
            if not first_ok:  # a wrong first attempt, then a correct retry
                add("answer", {**common, "correct": False, "attempt": 1,
                               "picked": random.choice(EMOTIONS),
                               "latencyMs": max(500, int(random.gauss(lat, 400)))})
                add("answer", {**common, "correct": True, "attempt": 2,
                               "picked": common["answer"],
                               "latencyMs": max(400, int(random.gauss(lat - 200, 300)))}, ev_score=score)
            else:
                score += 1
                p = {**common, "correct": True, "attempt": 1, "picked": common["answer"],
                     "latencyMs": max(400, int(random.gauss(lat, 350)))}
                if vr:
                    p["latencyFromPromptEndMs"] = max(250, int(random.gauss(lat - 600, 300)))
                    p.update(head_block(random.choice([-30, 30])))
                add("answer", p, ev_score=score)

    elif game in ("blocks", "playroom360"):
        for k in range(n):
            if random.random() < acc:  # placed on their turn
                score += 1
                p = {"round": k // 3 + 1, "slot": k % 3, "method": "tap" if vr else "button"}
                if vr:
                    p.update(head_block(random.choice([-25, 0, 25])))
                add("place_block", p, ev_score=score)
            else:  # jumped in out of turn
                add("impatient_tap", {"round": k // 3 + 1, "during": "peer-turn", "source": "tap"})

    elif game in ("rollback", "football360"):
        for k in range(n):
            cue = random.choice(CUES)
            if random.random() < 0.12:
                add("premature_roll", {"picked": f"p{random.randint(1,3)}", "cue": cue})
            first = random.random() < acc
            score += 1 if first else 0
            p = {"correct": True, "target": "p2", "picked": "p2", "cue": cue,
                 "firstAttempt": first, "latencyMs": max(500, int(random.gauss(lat, 400)))}
            if vr:
                b = random.choice([-80, -50, 50, 80])
                p.update({"targetBearingDeg": b, "latencyFromPromptEndMs": max(250, int(random.gauss(lat - 700, 300)))})
                p.update(head_block(b))
            add("roll_return", p, ev_score=score)

    elif game in ("rightway", "rulefixer", "rightway360"):
        cons = CONSTRUCTS.get(game, CONSTRUCTS["rightway"])
        for k in range(n):
            level = random.choice(LEVELS)
            correct = random.random() < acc
            score += 1 if correct else 0
            p = {"construct": random.choice(cons), "correct": correct, "level": level,
                 "chance": 0.5, "latencyMs": max(500, int(random.gauss(lat, 400)))}
            if game == "rulefixer":
                p["picked"] = f"opt{random.randint(1,3)}"
            add("answer", p, ev_score=score)

    elif game in ("museum", "museum360"):
        for k in range(n):
            cue, cuekind, vis = random.choice(MUSEUM_CUES)
            add("cue_ready", {"target": f"ex{random.randint(1,vis)}", "cue": cue, "cueKind": cuekind})
            first = random.random() < acc
            if not first:  # a wrong tap, then the correct one
                add("answer", {"correct": False, "target": "ex1", "picked": "ex2", "cue": cue,
                               "cueKind": cuekind, "visibleCount": vis,
                               "latencyMs": max(600, int(random.gauss(lat, 400))), "errorType": "adjacent"})
            score += 1 if first else 0
            p = {"correct": True, "target": "ex1", "picked": "ex1", "cue": cue, "cueKind": cuekind,
                 "visibleCount": vis, "firstAttempt": first,
                 "latencyMs": max(500, int(random.gauss(lat, 400))), "found": k + 1}
            if game == "museum360":
                b = random.choice([-60, -30, 30, 60])
                p["targetBearingDeg"] = b
                p.update(head_block(b))
            add("answer", p, ev_score=score)

    elif game in ("discovery", "park360"):
        for k in range(n):
            disc = random.choice(["butterfly", "rainbow", "puppy", "kite"])
            add("event_ready", {"discovery": disc, "saliency": random.choice(["big", "subtle"])})
            spontaneous = random.random() < acc
            if not spontaneous:
                add("nudge", {"kind": "share", "count": 1})
            score += 1 if spontaneous else 0
            p = {"correct": spontaneous, "spontaneous": spontaneous,
                 "nudges": 0 if spontaneous else random.randint(1, 2),
                 "latencyMs": max(800, int(random.gauss(lat + 400, 600))), "discovery": disc}
            if game == "park360":
                b = random.choice([-45, 0, 45])
                p["targetBearingDeg"] = b
                p.update(head_block(b))
            add("share", p, ev_score=score)

    add("game_over", {}, ev_score=score, dt_ms=step * 4200 + 1500)
    return evs, score


# --- build the study ----------------------------------------------------------
all_events = []          # flat list of Ev
sessions = []            # SimpleNamespace(student_id, game_key, started_at, ended_at)
events_by_student = {}   # student_id -> list[Ev]
roster = []              # participant dicts

for code, gender, dob, level, iq, profile in PARTICIPANTS:
    sid = uuid.uuid4()
    uid = uuid.uuid4()
    roster.append(dict(code=code, sid=sid, gender=gender, dob=dob, level=level,
                       iq=iq, profile=profile))
    events_by_student[sid] = []

    # latent ability scales with IQ; social-emotional baseline 0.45–0.8
    ability = clamp(0.45 + (iq - 60) / 100.0, 0.4, 0.82)
    games = list(PROFILE_GAMES[profile])
    smin, smax = PROFILE_SESSIONS[profile]

    # schedule: each game gets a few sessions; spread play across calendar days
    day_cursor = 0
    for game in games:
        n_sess = random.randint(smin, smax)
        for si in range(n_sess):
            # learning: accuracy rises across a game's sessions; latency falls
            acc = clamp(ability + 0.05 * si + random.uniform(-0.06, 0.06))
            lat = int(max(700, random.gauss(2600 - 180 * si - (iq - 85) * 6, 250)))
            day_cursor += random.choice([0, 1, 2, 2, 3])  # sometimes same day, else gap
            start = STUDY_START + timedelta(days=day_cursor, minutes=random.randint(0, 240))
            sess_id = f"{code}-{game[:5]}-{si+1}"
            evs, final = make_events(game, sess_id, sid, uid, start, acc, lat)
            all_events.extend(evs)
            events_by_student[sid].extend(evs)
            sessions.append(SimpleNamespace(
                student_id=sid, game_key=game, started_at=start,
                ended_at=start + timedelta(minutes=random.randint(4, 12)),
            ))

# --- sheet: participants ------------------------------------------------------
today = date(2026, 7, 18)
participants_rows = [[
    r["code"], str(r["sid"]), r["gender"], age_years(r["dob"], today),
    age_band(r["dob"], today), r["dob"].isoformat(), r["level"], r["iq"],
    iq_band(r["iq"]), r["profile"],
] for r in roster]
PARTICIPANT_COLS = ["participant_code", "student_id", "gender", "age_years", "age_band",
                    "date_of_birth", "autism_level", "iq_score", "iq_band", "profile_note"]

# --- sheet: raw_events (matches /export/events_raw.csv) -----------------------
RAW_FIXED = ["event_id", "participant_code", "student_id", "game_key", "event_type",
             "step_index", "event_score", "created_at", "gender", "date_of_birth",
             "autism_level", "iq_score"]
code_by_sid = {r["sid"]: r["code"] for r in roster}
demo_by_sid = {r["sid"]: r for r in roster}
payload_cols = scoring.raw_payload_columns(all_events)
RAW_COLS = RAW_FIXED + payload_cols


def raw_cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return int(v)  # 1/0 is friendlier for SPSS than True/False
    return v


all_events.sort(key=lambda e: (code_by_sid[e.student_id], e.created_at))
raw_rows = []
for e in all_events:
    r = demo_by_sid[e.student_id]
    p = e.payload
    raw_rows.append(
        [str(e.id), r["code"], str(e.student_id), e.game_key, e.event_type,
         e.step_index, "" if e.score is None else e.score, e.created_at.isoformat(),
         r["gender"], r["dob"].isoformat(), r["level"], r["iq"]]
        + [raw_cell(p.get(k)) for k in payload_cols]
    )

# --- sheet: trials (matches /export/trials.csv via student_trial_records) -----
TRIAL_COLS = ["participant_code", "student_id", "gender", "age_years", "age_band",
              "autism_level", "iq_score", "iq_band", "skill", "game_key", "xr_presenting",
              "session_id", "trial_in_game", "trial_in_session", "first_attempt_correct",
              "chance", "latency_ms", "latency_from_prompt_end_ms", "hinted", "construct",
              "cue", "visible_count", "head_yaw_travel_deg", "head_yaw_range_deg",
              "head_reversals", "head_to_target_ms", "timestamp"]
trial_rows = []
for r in roster:
    demo = [r["code"], str(r["sid"]), r["gender"], age_years(r["dob"], today),
            age_band(r["dob"], today), r["level"], r["iq"], iq_band(r["iq"])]
    for tr in scoring.student_trial_records(events_by_student[r["sid"]]):
        trial_rows.append(demo + [
            tr.skill, tr.game_key, "" if tr.xr_presenting is None else tr.xr_presenting,
            tr.session_id or "", tr.trial_in_game, tr.trial_in_session,
            tr.first_attempt_correct, tr.chance,
            "" if tr.latency_ms is None else tr.latency_ms,
            "" if tr.latency_from_prompt_end_ms is None else tr.latency_from_prompt_end_ms,
            "" if tr.hinted is None else tr.hinted, tr.construct, tr.cue,
            "" if tr.visible_count is None else tr.visible_count,
            "" if tr.head_yaw_travel_deg is None else tr.head_yaw_travel_deg,
            "" if tr.head_yaw_range_deg is None else tr.head_yaw_range_deg,
            "" if tr.head_reversals is None else tr.head_reversals,
            "" if tr.head_to_target_ms is None else tr.head_to_target_ms,
            tr.ts.isoformat(),
        ])

# --- sheet: dose (matches /export/dose.csv via dose_summary) ------------------
DOSE_COLS = ["participant_code", "student_id", "gender", "age_years", "age_band",
             "autism_level", "iq_score", "iq_band", "skill", "game_key", "n_sessions",
             "n_scored_trials", "total_minutes", "first_session", "last_session",
             "span_days", "median_gap_days"]
dose_rows = []
for r in roster:
    demo = [r["code"], str(r["sid"]), r["gender"], age_years(r["dob"], today),
            age_band(r["dob"], today), r["level"], r["iq"], iq_band(r["iq"])]
    trials_by_game = {}
    for tr in scoring.student_trial_records(events_by_student[r["sid"]]):
        trials_by_game[tr.game_key] = trials_by_game.get(tr.game_key, 0) + 1
    spans_by_game = {}
    for s in sessions:
        if s.student_id == r["sid"] and s.game_key in scoring.SKILL_BY_GAME:
            spans_by_game.setdefault(s.game_key, []).append((s.started_at, s.ended_at))
    for game in sorted(set(trials_by_game) | set(spans_by_game)):
        d = scoring.dose_summary(spans_by_game.get(game, []))
        dose_rows.append(demo + [
            scoring.SKILL_BY_GAME.get(game, ""), game, d.n_sessions,
            trials_by_game.get(game, 0), "" if d.total_minutes is None else d.total_minutes,
            d.first_session.date().isoformat() if d.first_session else "",
            d.last_session.date().isoformat() if d.last_session else "",
            "" if d.span_days is None else d.span_days,
            "" if d.median_gap_days is None else d.median_gap_days,
        ])

# --- sheet: battery (blinded pre/post outcomes) -------------------------------
BATTERY_COLS = ["participant_code", "timepoint", "instrument", "form", "raw_score",
                "n_options", "max_score", "rater_id", "is_double_coded", "assessed_on", "notes"]
# (instrument, max, n_options, gains for a completer, control?)
INSTRUMENTS = [
    ("EIT", 30, 3, 7, False),   # emotion identification, forced choice
    ("TOP", 21, None, 5, False),  # turn-taking observation
    ("JAP", 16, None, 4, False),  # joint attention probe
    ("NCT", 12, 4, 0, True),    # non-social control — should NOT improve
    ("VSMS", None, None, 4, False),  # Vineland social quotient (distal)
    ("ATEC", None, None, -6, False),  # ATEC total, lower = better (distal)
]
PRE_DAY = date(2026, 5, 28)
POST_DAY = date(2026, 7, 15)
battery_rows = []
for r in roster:
    ability = clamp(0.45 + (r["iq"] - 60) / 100.0, 0.4, 0.82)
    has_post = r["profile"] != "dropout"  # dropouts miss the post battery
    for inst, maxv, nopt, gain, control in INSTRUMENTS:
        form_pre, form_post = ("A", "B") if r["iq"] % 2 == 0 else ("B", "A")
        if inst in ("EIT", "TOP", "JAP", "NCT"):
            pre = int(round((maxv) * clamp(ability - 0.1, 0.3, 0.85)))
        elif inst == "VSMS":
            pre = int(round(55 + ability * 35))
        else:  # ATEC total 20–100, higher = more symptoms
            pre = int(round(90 - ability * 45))
        battery_rows.append([r["code"], "pre", inst, form_pre, pre, nopt or "",
                             maxv or "", "R1", "false", PRE_DAY.isoformat(), ""])
        if has_post:
            noise = random.randint(-1, 1)
            post = pre + (0 if control else gain) + noise
            if maxv:
                post = min(maxv, max(0, post))
            battery_rows.append([r["code"], "post", inst, form_post, post, nopt or "",
                                 maxv or "", "R1", "false", POST_DAY.isoformat(), ""])
    # one double-coded EIT-pre per completer (inter-rater reliability illustration)
    if r["profile"] == "completer":
        base = next(b for b in battery_rows if b[0] == r["code"] and b[1] == "pre" and b[2] == "EIT")
        battery_rows.append([r["code"], "pre", "EIT", base[3], base[4] + random.choice([-1, 0, 1]),
                             3, 30, "R2", "true", PRE_DAY.isoformat(), "second blinded coder"])

# --- sheet: summary (ONE row per participant — analysis-ready wide format) ----
# In-game skill scores (via the real score_participant) + dose totals + battery
# pre/post/gain, all on one row: the classic between-subjects SPSS layout.
SKILL_ORDER = ["emotion", "turntaking", "socialnorms", "jointattention"]
INSTR_ORDER = ["EIT", "TOP", "JAP", "NCT", "VSMS", "ATEC"]


def mean_nn(vals):
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 1) if vals else ""


# battery lookup, primary rater only (R2 double-codings excluded here)
batt_lookup = {}
for row in battery_rows:
    code, tp, inst, form, raw, nopt, maxv, rater, dbl, adate, notes = row
    if rater != "R1":
        continue
    batt_lookup.setdefault(code, {}).setdefault(inst, {})[tp] = raw

SUMMARY_COLS = ["participant_code", "student_id", "gender", "age_years", "age_band",
                "autism_level", "iq_score", "iq_band", "profile_note",
                "composite_score", "composite_delta"]
for sk in SKILL_ORDER:
    SUMMARY_COLS += [f"{sk}_pre", f"{sk}_post", f"{sk}_delta"]  # in-game, 0-100 chance-corrected
SUMMARY_COLS += ["total_sessions", "total_scored_trials", "total_minutes", "active_days",
                 "n_games_played", "n_skills_covered", "has_post_battery"]
for inst in INSTR_ORDER:
    i = inst.lower()
    SUMMARY_COLS += [f"{i}_pre", f"{i}_post", f"{i}_gain"]  # blinded outcome battery

summary_rows = []
for r in roster:
    sid, code = r["sid"], r["code"]
    ps = scoring.score_participant(events_by_student[sid])
    skill_by = {s.skill: s for s in ps.skills}

    my_sessions = [s for s in sessions if s.student_id == sid]
    total_minutes = round(sum((s.ended_at - s.started_at).total_seconds()
                              for s in my_sessions) / 60.0, 1) if my_sessions else 0.0
    starts = sorted(s.started_at for s in my_sessions)
    active_days = (starts[-1].date() - starts[0].date()).days if len(starts) >= 2 else 0
    trecs = scoring.student_trial_records(events_by_student[sid])
    has_post = 1 if any(batt_lookup.get(code, {}).get(i, {}).get("post") is not None
                        for i in INSTR_ORDER) else 0

    row = [code, str(sid), r["gender"], age_years(r["dob"], today), age_band(r["dob"], today),
           r["level"], r["iq"], iq_band(r["iq"]), r["profile"],
           "" if ps.composite is None else ps.composite,
           "" if ps.composite_delta is None else ps.composite_delta]
    for sk in SKILL_ORDER:
        s = skill_by.get(sk)
        if s and s.score is not None:
            scored = [g for g in s.games if g.score is not None]
            row += [mean_nn([g.baseline_score for g in scored]),
                    mean_nn([g.latest_score for g in scored]),
                    "" if s.delta is None else s.delta]
        else:
            row += ["", "", ""]  # skill never played (real missingness)
    row += [len(my_sessions), len(trecs), total_minutes, active_days,
            len({s.game_key for s in my_sessions}), len({t.skill for t in trecs}), has_post]
    for inst in INSTR_ORDER:
        pre = batt_lookup.get(code, {}).get(inst, {}).get("pre", "")
        post = batt_lookup.get(code, {}).get(inst, {}).get("post", "")
        gain = (int(post) - int(pre)) if (pre != "" and post != "") else ""
        row += [pre, post, gain]
    summary_rows.append(row)

# --- write CSVs + XLSX --------------------------------------------------------
SHEETS = [
    ("participants", PARTICIPANT_COLS, participants_rows),
    ("summary", SUMMARY_COLS, summary_rows),
    ("trials", TRIAL_COLS, trial_rows),
    ("dose", DOSE_COLS, dose_rows),
    ("battery", BATTERY_COLS, battery_rows),
    ("raw_events", RAW_COLS, raw_rows),
]
for name, cols, rows in SHEETS:
    with open(os.path.join(OUT, f"{name}.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        w.writerows(rows)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

# README sheet
readme = wb.create_sheet("README")
readme_lines = [
    ("Autism Games — DUMMY research dataset (synthetic, for format preview only)", True),
    ("", False),
    (f"{len(roster)} participants · {len(sessions)} game sessions · {len(all_events)} events · generated 2026-07-18", False),
    ("", False),
    ("Sheets:", True),
    ("  participants  — one row per child: code, demographics, profile_note (completer/partial/dropout)", False),
    ("  summary       — ONE row per child, analysis-ready wide format: in-game skill scores (pre/post/delta, 0-100), dose totals, battery pre/post/gain. Start here for between-subjects analysis.", False),
    ("  trials        — one row per SCORED trial (= /export/trials.csv). first_attempt_correct 0/1, chance, latency, VR fields.", False),
    ("  dose          — one row per participant × game: sessions, trials, minutes, spacing (= /export/dose.csv).", False),
    ("  battery        — blinded pre/post outcomes EIT/TOP/JAP/NCT + VSMS/ATEC (= the AssessmentScore import format).", False),
    ("  raw_events    — one row per recorded EVENT, all payload fields flattened (= /export/events_raw.csv). Source data for SPSS.", False),
    ("", False),
    ("Notes for analysis:", True),
    ("  · Missingness is intentional: dropouts (P-005) have few sessions and NO post battery; partials skip whole skills.", False),
    ("  · Accuracy rises and latency falls across a game's repeated sessions (a learning curve to detect).", False),
    ("  · NCT is the discriminant control — it should NOT improve pre→post, unlike the trained constructs.", False),
    ("  · xr_presenting = 1 rows are VR sessions and carry head-scan telemetry; Schoolyard 360 is flat (0).", False),
    ("  · Chance-correct accuracy yourself:  score = 100*max(0,(p-chance)/(1-chance)).", False),
    ("  · All values are FAKE. Structure mirrors the real exports exactly.", False),
]
for i, (txt, bold) in enumerate(readme_lines, start=1):
    c = readme.cell(row=i, column=1, value=txt)
    c.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
readme.column_dimensions["A"].width = 120

HEAD_FILL = PatternFill("solid", fgColor="1E293B")
HEAD_FONT = Font(bold=True, color="E2E8F0")
for name, cols, rows in SHEETS:
    ws = wb.create_sheet(name)
    ws.append(cols)
    for r in rows:
        ws.append(r)
    for ci, _ in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="left")
        letter = get_column_letter(ci)
        width = min(26, max(11, max([len(str(cols[ci - 1]))] + [len(str(r[ci - 1])) for r in rows[:200]] or [10]) + 2))
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

xlsx_path = os.path.join(OUT, "dummy_research_data.xlsx")
wb.save(xlsx_path)

print("participants:", len(roster))
print("sessions    :", len(sessions))
print("raw events  :", len(all_events), "| raw columns:", len(RAW_COLS))
print("trials      :", len(trial_rows))
print("dose rows   :", len(dose_rows))
print("battery rows:", len(battery_rows))
print("written to  :", OUT)
print("files       :", ", ".join(sorted(os.listdir(OUT))))
